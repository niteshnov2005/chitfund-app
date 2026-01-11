import openpyxl
import os

FILE_NAME = r"d:\chitfundproject\sample_gemini.xlsx"

if not os.path.exists(FILE_NAME):
    print("File not found")
    exit()

wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

TARGET_VAL = 8855450

print(f"Scanning {sheet_name} for {TARGET_VAL}...")

found = False
for r in range(1, 1000):
    for c in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=r, column=c).value
        try:
            # Check exact match
            if cell_val == TARGET_VAL:
                print(f"MATCH FOUND! Row {r}, Col {c} Value: {cell_val}")
                found = True
            
            # Check string representation match
            str_val = str(cell_val).replace(',', '').replace('₹', '').strip()
            if str_val == str(TARGET_VAL) or str_val == str(TARGET_VAL) + ".0":
                 print(f"STRING MATCH FOUND! Row {r}, Col {c} Value: {cell_val}")
                 found = True
                 
        except: pass

if not found:
    print("Value not found in first 1000 rows.")
    # Calculate actual sum of all numbers in 'Amount' columns to see what it is
    print("Calculating expected sum from known columns...")
