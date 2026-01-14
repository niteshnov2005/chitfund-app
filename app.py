from flask import Flask, render_template, jsonify, request, redirect, url_for, session
from datetime import datetime
import pandas as pd
import openpyxl
import os
import time
import traceback
import json

import io
from flask import Flask, render_template, jsonify, request, redirect, url_for, session, send_file

app = Flask(__name__)

app.secret_key = 'mini_project_secret_key'
app.secret_key = 'mini_project_secret_key'

# --- CONFIGURATION (Start) ---
# Detect environment
IS_CLOUD_RUN = os.environ.get('K_SERVICE') is not None
HAS_SUPABASE = os.environ.get("SUPABASE_URL") is not None

if IS_CLOUD_RUN:
    # Cloud Run (Stateless): Must use /tmp
    FILE_NAME = '/tmp/data.xlsx'
else:
    # Local Windows or PythonAnywhere (Stateful): Use local file
    # Ensure absolute path for PythonAnywhere robustness
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    FILE_NAME = os.path.join(BASE_DIR, 'sample_gemini.xlsx')

USE_CLOUD_STORAGE = HAS_SUPABASE
BUCKET_NAME = os.environ.get("BUCKET_NAME", "chitfund-data")

# --- CLOUD STORAGE HELPERS (SUPABASE) ---
def sync_down():
    """Downloads the excel file from Supabase Storage if it exists."""
    if not USE_CLOUD_STORAGE: return
    try:
        from supabase import create_client
        url = os.environ.get("SUPABASE_URL")
        key = os.environ.get("SUPABASE_KEY")
        if not url or not key:
            print("Supabase credentials missing.")
            return

        supabase = create_client(url, key)
        
        # Download file (returns bytes)
        print(f"Downloading data.xlsx from bucket '{BUCKET_NAME}'...")
        response = supabase.storage.from_(BUCKET_NAME).download('data.xlsx')
        
        if response:
            with open(FILE_NAME, 'wb') as f:
                f.write(response)
            print("Download successful.")
        else:
            print("No remote file found.")
            # If Cloud Run and no remote file, we might crash if we don't have a seed.
            # But currently we assume seed is in bucket.
            
    except Exception as e:
        print(f"Supabase Download Error: {e}")

# Initial Sync
if USE_CLOUD_STORAGE: sync_down()

# Cloud Run Fallback: If /tmp/data.xlsx still doesn't exist (sync failed or new setup),
# try to copy from local seed if available in container
if IS_CLOUD_RUN and not os.path.exists(FILE_NAME):
    local_seed = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sample_gemini.xlsx')
    if os.path.exists(local_seed):
        print("Copying local seed to /tmp...")
        import shutil
        shutil.copy(local_seed, FILE_NAME)

def sync_up():
    """Uploads the local excel file to Supabase Storage."""
    if not USE_CLOUD_STORAGE: return
    try:
        from supabase import create_client
        url = os.environ.get("SUPABASE_URL")
        key = os.environ.get("SUPABASE_KEY")
        if not url or not key: return

        supabase = create_client(url, key)
        
        print(f"Uploading data.xlsx to bucket '{BUCKET_NAME}'...")
        with open(FILE_NAME, 'rb') as f:
            # Upsert is handled via file_options in some versions, or explicit 'upload' with overwriting behavior
            # supabase-py storage upload signature varies. safer to try upload, if fail try update (or use upsert option)
            # Standard py client usually supports file_options={"upsert": "true"}
            supabase.storage.from_(BUCKET_NAME).upload('data.xlsx', f, file_options={"upsert": "true"})
            
        print("Upload successful.")
    except Exception as e:
        # Fallback: if upload fails (maybe file exists and upsert didn't work), try 'update'
        try:
           print(f"Upload failed ({e}), trying update...")
           with open(FILE_NAME, 'rb') as f:
               supabase.storage.from_(BUCKET_NAME).update('data.xlsx', f, file_options={"upsert": "true"})
           print("Update successful.")
        except Exception as e2:
           print(f"Supabase Upload/Update Error: {e2}")

# Initial Sync
if USE_CLOUD_STORAGE: sync_down()

# --- UTILS ---
def clean_num(val):
    try:
        return float(str(val).replace(',', '').replace('₹', '').strip())
    except:
        return 0.0

def clean_plan_amount(plan_str):
    try:
        cleaned = str(plan_str).replace(',', '').replace('₹', '').split('.')[0]
        return int(cleaned)
    except: return 0

def normalize_text(text):
    if not text: return ""
    return str(text).lower().replace('.', '').replace(' ', '').replace(':', '').strip()

# --- PAYMENT RECORDS DB HANDLER ---
PAID_DB_FILE = 'payment_records.json'

def load_paid_db():
    if not os.path.exists(PAID_DB_FILE): return {}
    try:
        with open(PAID_DB_FILE, 'r') as f: return json.load(f)
    except: return {}

def save_paid_db(data):
    with open(PAID_DB_FILE, 'w') as f: json.dump(data, f, indent=4)

# --- HELPER: GET ALL SHEET NAMES (MONTHS) ---
def get_all_sheet_names():
    if not os.path.exists(FILE_NAME): return []
    try:
        wb = openpyxl.load_workbook(FILE_NAME, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except: return []

# --- 1. DASHBOARD READER (Supports specific sheet_name) ---
def get_excel_data(sheet_name=None):
    if not os.path.exists(FILE_NAME): return []
    try:
        # If no sheet specified, default to the LAST sheet (Most recent month)
        if not sheet_name:
            wb = openpyxl.load_workbook(FILE_NAME, read_only=True)
            sheet_name = wb.sheetnames[-1] # Assume last sheet is latest
            wb.close()
            
        df = pd.read_excel(FILE_NAME, sheet_name=sheet_name, header=None, engine='openpyxl')
    except: return []

    paid_db = load_paid_db()
    members_list = []

    # --- STEP 1: BUILD MASTER LIST ---
    member_map = {}
    name_only_map = {}

    summary_blocks = [
        {'amt': 19, 'name': 20, 'area': 21}, # Cols T, U, V
        {'amt': 23, 'name': 24, 'area': 25}  # Cols X, Y, Z
    ]
    exclude_values = [4132350.0, 3708850.0, 7841200.0]

    for block in summary_blocks:
        for r in range(df.shape[0]):
            try:
                raw_name = str(df.iloc[r, block['name']]).strip().title()
                if raw_name.upper() in ['NAN', 'NAME', 'AMOUNT', 'TOTAL', '', '0', 'TOTAL PAYABLE', 'AREA']: continue

                amt = clean_num(df.iloc[r, block['amt']])
                if any(abs(amt - ex) < 1 for ex in exclude_values): continue

                raw_area = "General"
                if block['area'] < df.shape[1]:
                    val = str(df.iloc[r, block['area']]).strip().title()
                    if val and val.upper() != 'NAN' and not val.replace('.','',1).isdigit():
                        raw_area = val

                # Create Keys
                norm_n = normalize_text(raw_name)
                norm_a = normalize_text(raw_area)
                unique_key = f"{norm_n}_{norm_a}"

                # --- DUPLICATE CHECK & MERGE ---
                if unique_key in member_map:
                    existing = member_map[unique_key]
                    existing['total'] += amt

                    new_id = f"{existing['name']}_{int(existing['total'])}".replace(" ", "")
                    existing['payment_id'] = new_id

                    db_entry = paid_db.get(new_id, None)
                    if db_entry:
                        if db_entry is True:
                            existing['is_paid'] = True
                            existing['paid_date'] = None
                        elif isinstance(db_entry, str):
                            existing['is_paid'] = True
                            existing['paid_date'] = db_entry
                    else:
                        existing['is_paid'] = False
                        existing['paid_date'] = None

                else:
                    payment_id = f"{raw_name}_{int(amt)}".replace(" ", "")

                    db_entry = paid_db.get(payment_id, None)
                    is_paid = False
                    paid_date = None

                    if db_entry:
                        if db_entry is True:
                            is_paid = True
                        elif isinstance(db_entry, str):
                            is_paid = True
                            paid_date = db_entry

                    member_obj = {
                        'name': raw_name,
                        'area': raw_area,
                        'total': amt,
                        'items': [],
                        'payment_id': payment_id,
                        'is_paid': is_paid,
                        'paid_date': paid_date
                    }

                    members_list.append(member_obj)
                    member_map[unique_key] = member_obj

                    if norm_n not in name_only_map: name_only_map[norm_n] = []
                    name_only_map[norm_n].append(member_obj)

            except: pass

    # --- STEP 2: SCAN RECEIPTS ---
    receipt_scan_configs = [
        {'col_idx': 0, 'name_col': 1, 'area_col': 2, 'amt_col': 3},
        {'col_idx': 5, 'name_col': 6, 'area_col': 7, 'amt_col': 8},
        {'col_idx': 10, 'name_col': 11, 'area_col': 12, 'amt_col': 13} # Added likely 3rd column block
    ]

    for config in receipt_scan_configs:
        col_start = config['col_idx']
        name_col_idx = config['name_col']
        area_col_idx = config['area_col']

        curr_name = None
        curr_area = "General"

        for r in range(df.shape[0]):
            try:
                c0 = str(df.iloc[r, col_start]).strip()
                if 'NAME' in c0.upper() and len(c0) < 25:
                    curr_name = str(df.iloc[r, name_col_idx]).strip().title()
                    # Validation: Skip empty names
                    if not curr_name or curr_name.upper() in ['NONE', 'NAN', '']:
                        curr_name = None
                        continue

                    val = str(df.iloc[r, area_col_idx]).strip().title()
                    if val and val.upper() != 'NAN' and not val.replace('.','',1).isdigit():
                        curr_area = val
                    else:
                        curr_area = "General"
                    
                    # --- IMMEDIATE MEMBER CREATION ---
                    n_receipt = normalize_text(curr_name)
                    a_receipt = normalize_text(curr_area)
                    key = f"{n_receipt}_{a_receipt}"
                    
                    target_member = None
                    if key in member_map:
                        target_member = member_map[key]
                    else:
                        # Fuzzy search fallback (check existing list)
                        candidates = name_only_map.get(n_receipt)
                        if candidates: # logic to find cand...
                             if len(candidates) == 1: target_member = candidates[0]
                             # simplify fuzzy for header creation?
                        
                    if not target_member:
                        new_id = f"{key}_auto"
                        target_member = {
                            'name': curr_name,
                            'area': curr_area,
                            'total': 0,
                            'paid_amount': 0,
                            'items': [],
                            'payment_id': new_id,
                            'is_paid': False
                        }
                        members_list.append(target_member)
                        member_map[key] = target_member
                        if n_receipt not in name_only_map: name_only_map[n_receipt] = []
                        name_only_map[n_receipt].append(target_member)

                if curr_name and c0.replace('.','',1).isdigit():
                    # Check payment amounts...
                    # We need to re-fetch target_member here properly to add items
                    
                    curr_member_obj = target_member # Set state

                # --- ITEM DETECTION ---
                if curr_member_obj and c0.replace('.','',1).isdigit():
                    amt = clean_num(df.iloc[r, config['amt_col']])
                    if amt > 0:
                        n_receipt = normalize_text(curr_member_obj['name']) # Use obj name
                        
                        month = str(df.iloc[r, col_start]).strip()
                        plan = str(df.iloc[r, col_start+1]).strip()
                        commission = str(df.iloc[r, col_start+2]).strip()

                        # Generate Granular Item ID common for both branches
                        safe_month = month.replace(" ", "")
                        safe_plan = plan.replace(" ", "")
                        item_id = f"{n_receipt}_{int(amt)}_{safe_month}_{safe_plan}".replace(" ", "")

                        # Check if this specific item is paid
                        db_entry = paid_db.get(item_id)
                        item_is_paid = False
                        item_paid_date = None
                        
                        if db_entry:
                            if db_entry is True: item_is_paid = True
                            elif isinstance(db_entry, str): 
                                item_is_paid = True
                                item_paid_date = db_entry

                        # Backward Compatibility Check
                        if not item_is_paid and paid_db.get(curr_member_obj['payment_id']):
                            db_val = paid_db.get(curr_member_obj['payment_id'])
                            item_is_paid = True
                            item_paid_date = db_val if isinstance(db_val, str) else None

                        curr_member_obj['items'].append({
                            'month': month,
                            'plan': plan,
                            'commission': commission,
                            'amount': amt,
                            'id': item_id,
                            'is_paid': item_is_paid,
                            'paid_date': item_paid_date
                        })
            except: pass

    # --- STEP 3: ATTACH COMMISSIONS ---
    comm_col = -1
    for c in range(df.shape[1]):
        for r in range(20):
            if 'TOTAL COMMISSION' in str(df.iloc[r, c]).upper(): comm_col = c; break

    if comm_col != -1:
        for r in range(df.shape[0]):
            val = clean_num(df.iloc[r, comm_col])
            if val > 0:
                if 'TOTAL' in str(df.iloc[r, comm_col-1]).upper(): continue
                name_guess = str(df.iloc[r, 1]).strip().title()
                if name_guess:
                     norm_n = normalize_text(name_guess)
                     candidates = name_only_map.get(norm_n)
                     if candidates and len(candidates) > 0:
                         candidates[0]['items'].append({
                             'month': '-', 
                             'plan': 'Comm.', 
                             'commission': '-', 
                             'amount': val,
                             'id': f"comm_{int(val)}_{r}", 
                             'is_paid': True, # Commissions are effectively "paid" or credits? Actually let's assume auto-paid/credit.
                             'paid_date': None
                         })

    # --- STEP 4: RE-CALCULATE MEMBER TOTALS with Partial Logic ---
    final_list = []
    
    for m in members_list:
        if m['items']:
            # Calculate total from items to be safe
            m['total'] = sum(i['amount'] for i in m['items'])
            m['paid_amount'] = sum(i['amount'] for i in m['items'] if i['is_paid'])
            m['is_paid'] = (m['paid_amount'] >= m['total']) and (m['total'] > 0)
        else:
            # Fallback for members without items
            m['paid_amount'] = m['total'] if m['is_paid'] else 0
            
        final_list.append(m)

    # --- STEP 5: GET EXPLICIT GRAND TOTAL ---
    # The excel has a trusted Grand Total at Row 23, Col 15 (Index 22, 14 0-indexed)
    grand_total_val = 0
    try:
        # Check specific cell
        grand_total_val = clean_num(df.iloc[22, 14])
        # Validate order of magnitude (should be > 100k)
        if grand_total_val < 100000:
             # Fallback scan
             pass
    except: pass
    
    # Return structure with metadata
    return {
        'members': sorted(final_list, key=lambda x: x['name']),
        'grand_total': grand_total_val
    }

# --- 2. AUCTION READER ---
def get_auction_plans():
    if not os.path.exists(FILE_NAME): return []
    try: df = pd.read_excel(FILE_NAME, header=None, engine='openpyxl')
    except: return []
    plans = []
    for r in range(2, 50):
        try:
            m_val = str(df.iloc[r, 11]).strip()
            p_val_raw = str(df.iloc[r, 12]).strip()
            if not m_val or m_val.upper() == 'NAN': continue
            if 'TOTAL' in str(df.iloc[r, 13]).upper(): break
            p_val = clean_plan_amount(p_val_raw)
            if p_val > 0:
                plans.append({'current_month': m_val, 'plan_display': p_val_raw, 'plan_value': p_val, 'id': f"{m_val}_{p_val}", 'member_count': "-"})
        except: pass
    return plans

# --- 3. AUCTION UPDATER (FIXED: Month Format + Zero Commission Total) ---
@app.route('/run-auction-batch', methods=['POST'])
def run_auction_batch():
    if 'user' not in session: return redirect(url_for('login_page'))
    try:
        global_date = request.form.get('global_date')
        sheet_name_suffix = request.form.get('sheet_name')

        updates = {}
        # 1. Parse Inputs
        for key, val in request.form.items():
            if key.startswith('bid_for_'):
                unique_id = key.replace('bid_for_', '')
                new_m = request.form.get(f'new_month_for_{unique_id}')
                bid = request.form.get(f'bid_for_{unique_id}')

                # Check if new_m and bid exist (strings "0" are True, empty strings are False)
                if new_m and bid is not None:
                    try:
                        total_bid = float(bid)
                        plan_val = int(unique_id.split('_')[1])
                        dividend = total_bid / 20
                        payable_amount = (plan_val - total_bid) / 20

                        updates[unique_id] = {
                            'new_month': str(new_m).strip(),
                            'new_dividend': int(dividend),
                            'new_payable': int(payable_amount),
                            'plan_val': plan_val,
                            'total_bid': total_bid,
                            'old_month_prefix': unique_id.split('_')[0]
                        }
                    except: continue

        if not updates: return "Error: No Valid Inputs"

        # Helper: Month Matcher
        def is_same_month(m1, m2):
            s1 = str(m1).strip()
            s2 = str(m2).strip()
            if s1 == s2: return True
            try: return abs(float(s1) - float(s2)) < 0.001
            except: return False

        # Load Workbooks
        wb_data = openpyxl.load_workbook(FILE_NAME, data_only=True)
        source_sheet_data = wb_data.active

        wb = openpyxl.load_workbook(FILE_NAME)
        source_sheet = wb.active
        target_sheet = wb.copy_worksheet(source_sheet)
        target_sheet.title = sheet_name_suffix
        wb.move_sheet(target_sheet, offset=-len(wb.sheetnames)+1)
        wb.active = target_sheet

        # Helper: Month Formatter (20 vs 20.0)
        def process_month_value(val):
            try:
                f_val = float(val)
                if f_val.is_integer(): return int(f_val)
                return f_val
            except: return val

        # Map Summary Locations
        summary_locs = {}
        for name_col in [20, 24]:
             for r in range(1, target_sheet.max_row + 1):
                 name = str(target_sheet.cell(row=r, column=name_col).value).strip().title()
                 if name and name.upper() not in ['NAN', 'NAME', 'AMOUNT']:
                     if name not in summary_locs: summary_locs[name] = []
                     summary_locs[name].append({'row': r, 'col': name_col-1})

        # --- UPDATE LOOP 1: Main Ledger ---
        for col_start in [1, 6]:
            curr_member = None
            block_sum = 0      # Sum of Payable Amount
            dividend_sum = 0   # Sum of Commission/Dividend

            for row in range(1, target_sheet.max_row + 1):
                cell_val = str(target_sheet.cell(row=row, column=col_start).value).upper()

                # Header Reset
                if 'NAME' in cell_val:
                    curr_member = str(target_sheet.cell(row=row, column=col_start+1).value).strip().title()
                    target_sheet.cell(row=row, column=col_start+3).value = global_date
                    block_sum = 0
                    dividend_sum = 0 # Reset dividend sum for new block

                # Data Row Processing
                val_check = str(target_sheet.cell(row=row, column=col_start).value).strip()
                if val_check.replace('.','',1).isdigit():
                    curr_m = val_check
                    curr_p = str(target_sheet.cell(row=row, column=col_start+1).value).strip()
                    curr_pv = clean_plan_amount(curr_p)

                    old_pay_val = source_sheet_data.cell(row=row, column=col_start+3).value
                    old_pay = clean_num(old_pay_val)
                    current_item_amt = old_pay

                    # Get existing dividend (for sum calculation if not updated)
                    old_div_val = source_sheet_data.cell(row=row, column=col_start+2).value
                    current_dividend = clean_num(old_div_val)

                    matched = None
                    for uid, upd in updates.items():
                        if upd['plan_val'] == curr_pv and is_same_month(upd['old_month_prefix'], curr_m):
                            matched = upd; break

                    if matched:
                        new_pay = matched['new_payable']
                        diff = new_pay - old_pay

                        fm = process_month_value(matched['new_month'])

                        target_sheet.cell(row=row, column=col_start).value = fm
                        target_sheet.cell(row=row, column=col_start+2).value = matched['new_dividend']
                        target_sheet.cell(row=row, column=col_start+3).value = new_pay

                        current_item_amt = new_pay
                        current_dividend = matched['new_dividend'] # Update dividend for sum

                        # Update Summary Area (Right side of sheet)
                        if curr_member and curr_member in summary_locs:
                            for loc in summary_locs[curr_member]:
                                old_sum_val = source_sheet_data.cell(row=loc['row'], column=loc['col']).value
                                old_sum = clean_num(old_sum_val)
                                target_sheet.cell(row=loc['row'], column=loc['col']).value = int(old_sum) + int(diff)

                    block_sum += current_item_amt
                    dividend_sum += current_dividend # Add to running total

                # Footer Total Update
                if 'TOTAL' in cell_val:
                    target_sheet.cell(row=row, column=col_start+3).value = int(block_sum)
                    # FIX: Explicitly update the Total Commission cell
                    target_sheet.cell(row=row, column=col_start+2).value = int(dividend_sum)
                    block_sum = 0
                    dividend_sum = 0

        # --- UPDATE LOOP 2: Auction List ---
        for r in range(2, 100):
            try:
                p_val_raw = str(source_sheet_data.cell(row=r, column=13).value)
                p_val = clean_plan_amount(p_val_raw)
                m_val = str(source_sheet_data.cell(row=r, column=12).value).strip()

                if p_val > 0:
                    for uid, upd in updates.items():
                        if upd['plan_val'] == p_val and is_same_month(upd['old_month_prefix'], m_val):
                             fm = process_month_value(upd['new_month'])
                             target_sheet.cell(row=r, column=12).value = fm
                             target_sheet.cell(row=r, column=14).value = upd['total_bid']
                             target_sheet.cell(row=r, column=16).value = upd['new_dividend']
                             target_sheet.cell(row=r, column=18).value = upd['new_payable']
            except: pass

        wb.save(FILE_NAME)
        wb_data.close()
        # SYNC TO CLOUD
        if USE_CLOUD_STORAGE: sync_up()
        return redirect(url_for('dashboard'))

    except PermissionError: return "ERROR: Close Excel file!"
    except Exception as e: return f"Error: {e}\n{traceback.format_exc()}"

# --- ROUTES ---
@app.route('/')
def login_page(): return render_template('login.html') if 'user' not in session else redirect(url_for('dashboard'))
@app.route('/login', methods=['POST'])
def login_logic():
    if request.form.get('username') == 'admin' and request.form.get('password') == 'nitesh2025': session['user'] = 'admin'; return redirect(url_for('dashboard'))
    return render_template('login.html', error="Invalid Credentials")
@app.route('/logout')
def logout(): session.pop('user', None); return redirect(url_for('login_page'))
@app.route('/dashboard')
def dashboard(): return render_template('dashboard.html', timestamp=time.time()) if 'user' in session else redirect(url_for('login_page'))
@app.route('/members')
def members_page(): return render_template('members.html') if 'user' in session else redirect(url_for('login_page'))
@app.route('/auction')
def auction_page(): return render_template('auction.html', groups=get_auction_plans()) if 'user' in session else redirect(url_for('login_page'))
@app.route('/api/members')
def get_members_api():
    # Optional: allow fetching data for a specific sheet (Month)
    sheet = request.args.get('sheet')
    return jsonify(get_excel_data(sheet)) if 'user' in session else (jsonify({"error": "Unauthorized"}), 401)

@app.route('/api/sheets')
def get_sheets_api():
    return jsonify({'sheets': get_all_sheet_names()}) if 'user' in session else (jsonify({"error": "Unauthorized"}), 401)

# --- EXCEL EDITOR API ---
@app.route('/api/sheet_data')
def get_sheet_data_api():
    if 'user' not in session: return jsonify({'error': 'Unauthorized'}), 401
    sheet_name = request.args.get('sheet')
    if not sheet_name: return jsonify({'error': 'No sheet name'}), 400
    
    try:
        wb = openpyxl.load_workbook(FILE_NAME, data_only=False) # Get formulas if possible? No, for editing we want values usually, or raw? Jspreadsheet edits values.
        # Actually data_only=False is better to edit formulas, but for this use case (simple data entry), values are fine. 
        # But if we write back values, we lose formulas.
        # Let's stick to data_only=False to read/write what is there.
        if sheet_name not in wb.sheetnames: return jsonify({'error': 'Sheet not found'}), 404
        ws = wb[sheet_name]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            # cell values can be None, datetime, etc. Convert to friendly formats
            clean_row = []
            for cell in row:
                if cell is None: clean_row.append("")
                elif isinstance(cell, datetime): clean_row.append(cell.strftime('%Y-%m-%d'))
                else: clean_row.append(str(cell))
            data.append(clean_row)
            
        wb.close()
        return jsonify({'data': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/save_sheet_data', methods=['POST'])
def save_sheet_data_api():
    if 'user' not in session: return jsonify({'error': 'Unauthorized'}), 401
    try:
        req = request.json
        sheet_name = req.get('sheet_name')
        data = req.get('data') # 2D array
        
        if not sheet_name or not data: return jsonify({'error': 'Missing data'}), 400
        
        wb = openpyxl.load_workbook(FILE_NAME)
        if sheet_name not in wb.sheetnames: return jsonify({'error': 'Sheet not found'}), 404
        ws = wb[sheet_name]
        
        # We only update cells provided in 'data'. 
        # CAUTION: 'data' from Jspreadsheet might be the whole sheet or specific changes.
        # Ideally we expect the whole sheet data back to rewrite it? 
        # Jspreadsheet .getData() returns the whole grid.
        
        for r_idx, row in enumerate(data):
            for c_idx, val in enumerate(row):
                # r_idx+1, c_idx+1 because openpyxl is 1-based
                # Try to preserve types if possible? Everything is str from JSON.
                # Try converting to int/float if it looks like one
                clean_val = val
                if isinstance(val, str):
                    val = val.strip()
                    if val.replace('.','',1).isdigit():
                        try:
                            if '.' in val: clean_val = float(val)
                            else: clean_val = int(val)
                        except: clean_val = val
                    elif val == "":
                        clean_val = None
                
                ws.cell(row=r_idx+1, column=c_idx+1, value=clean_val)
                
        wb.save(FILE_NAME)
        wb.close()
        # SYNC TO CLOUD
        if USE_CLOUD_STORAGE: sync_up()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/toggle-pay', methods=['POST'])
def toggle_pay():
    if 'user' not in session: return jsonify({'error': 'Unauthorized'}), 401
    item_id = request.json.get('id')
    if not item_id: return jsonify({'error': 'No ID'}), 400
    db = load_paid_db()
    current_entry = db.get(item_id)
    if current_entry:
        del db[item_id]
        new_status = False
        paid_on = None
    else:
        # Changed: Use local server time instead of UTC to fix user reported time mistake.
        now = datetime.now().strftime("%d %b, %I:%M %p")
        db[item_id] = now
        new_status = True
        paid_on = now
    save_paid_db(db)
    return jsonify({'success': True, 'new_status': new_status, 'paid_on': paid_on})

@app.route('/download_excel')
def download_excel():
    if 'user' not in session: return redirect(url_for('login_page'))
    
    # 1. Get processed data
    members = get_excel_data() 
    
    # 2. Create Workbook using openpyxl directly for formatting
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Receipts"
    
    # Styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # Light Grey
    bold_font = Font(bold=True)
    
    # Column Widths
    ws.column_dimensions['A'].width = 15 # Month / Label
    ws.column_dimensions['B'].width = 25 # Plan / Name
    ws.column_dimensions['C'].width = 25 # Comm / Area
    ws.column_dimensions['D'].width = 20 # Amount
    
    row_idx = 1
    
    for m in members:
        # Sort out date
        display_date = m.get('paid_date', datetime.now().strftime("%d-%b-%Y"))
        if not display_date: display_date = datetime.now().strftime("%d-%b-%Y")
        
        # --- ROW 1: Header Block [Name Label | Name | Area | Date] ---
        # Cell A: Name Label
        c1 = ws.cell(row=row_idx, column=1, value="Name")
        c1.font = bold_font
        c1.border = thin_border
        
        # Cell B: Name Value
        c2 = ws.cell(row=row_idx, column=2, value=m['name'])
        c2.font = bold_font
        c2.border = thin_border
        
        # Cell C: Area Value
        c3 = ws.cell(row=row_idx, column=3, value=m['area'])
        c3.font = bold_font
        c3.border = thin_border
        
        # Cell D: Date
        c4 = ws.cell(row=row_idx, column=4, value=display_date)
        c4.font = bold_font
        c4.alignment = Alignment(horizontal='right')
        c4.border = thin_border
        
        row_idx += 1
        
        # --- ROW 2: Column Headers [Month | Plan | Commission | Amount] ---
        headers = ["Month", "Plan", "Commission", "Amount"]
        for col_num, h_text in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_num, value=h_text)
            c.fill = header_fill
            c.border = thin_border
        row_idx += 1
        
        # --- ROW 3+: Items ---
        items_to_print = m['items'] if m['items'] else [{'month': '-', 'plan': '-', 'commission': '-', 'amount': m['total']}]
        
        for item in items_to_print:
            # Month
            ws.cell(row=row_idx, column=1, value=item.get('month', '-')).border = thin_border
            # Plan
            # Try to format as number if possible
            plan_val = item.get('plan', '-')
            try: plan_val = float(str(plan_val).replace(',','')) 
            except: pass
            c_plan = ws.cell(row=row_idx, column=2, value=plan_val)
            c_plan.border = thin_border
            if isinstance(plan_val, (int, float)): c_plan.number_format = '#,##0.00'
            
            # Commission
            comm_val = item.get('commission', '-')
            try: comm_val = float(str(comm_val).replace(',','')) 
            except: pass
            c_comm = ws.cell(row=row_idx, column=3, value=comm_val)
            c_comm.border = thin_border
            if isinstance(comm_val, (int, float)): c_comm.number_format = '#,##0.00'

            # Amount
            amt_val = item.get('amount', 0)
            c_amt = ws.cell(row=row_idx, column=4, value=amt_val)
            c_amt.border = thin_border
            c_amt.number_format = '#,##0.00'
            
            row_idx += 1
            
        # --- ROW N: Empty spacers (to match image style roughly) ---
        # Image shows some empty grey rows or just space? 
        # Image has 2 empty rows with grey background before Total. Let's add 1 empty row.
        for col_num in range(1, 5):
             ws.cell(row=row_idx, column=col_num).fill = header_fill
             ws.cell(row=row_idx+1, column=col_num).fill = header_fill
        row_idx += 2

        # --- ROW Footer: Total Payable ---
        c_foot_label = ws.cell(row=row_idx, column=1, value="Total Payable")
        c_foot_label.font = bold_font
        c_foot_label.fill = header_fill
        c_foot_label.border = thin_border
        
        # Merge B and C for footer or just leave empty? Image shows total at right.
        ws.cell(row=row_idx, column=2).fill = header_fill
        ws.cell(row=row_idx, column=2).border = thin_border
        ws.cell(row=row_idx, column=3).fill = header_fill
        ws.cell(row=row_idx, column=3).border = thin_border
        
        c_foot_val = ws.cell(row=row_idx, column=4, value=m['total'])
        c_foot_val.font = bold_font
        c_foot_val.fill = header_fill
        c_foot_val.border = thin_border
        c_foot_val.number_format = '#,##0.00'
        
        row_idx += 1
        
        # Spacer between members
        row_idx += 3

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Payment_Receipts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

from itertools import groupby

@app.route('/view_receipts')
def view_receipts():
    if 'user' not in session: return redirect(url_for('login_page'))
    
    # Default to the most recent sheet (Last one) if no sheet specified?
    # Actually, let's explicit pull the last sheet name.
    sheet_names = get_all_sheet_names()
    target_sheet = sheet_names[-1] if sheet_names else None
    
    data_response = get_excel_data(target_sheet)
    # Handle dict response
    if isinstance(data_response, dict):
        members = data_response.get('members', [])
    else:
        members = data_response
    
    # Sort by Area (primary) and Name (secondary)
    def sort_key(x):
        return (str(x.get('area', '')).strip().lower(), str(x.get('name', '')).strip().lower())
    
    members.sort(key=sort_key)
    
    # Continuous Layout: Return flat list, sorted by Area
    return render_template('receipt_preview.html', members=members, now=datetime.now())

# --- REPORT ROUTES ---
@app.route('/reports')
def reports_page():
    if 'user' not in session: return redirect(url_for('login_page'))
    sheets = get_all_sheet_names()
    return render_template('reports.html', sheets=sheets)

@app.route('/excel_editor')
def excel_editor_page():
    return render_template('excel_editor.html') if 'user' in session else redirect(url_for('login_page'))

if __name__ == '__main__':
    # Local Dev (Windows)
    app.run(debug=True)
else:
    # Production / WSGI (PythonAnywhere)
    application = app
