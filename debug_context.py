import openpyxl
import os

FILE_NAME = r"d:\chitfundproject\sample_gemini.xlsx"
wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
ws = wb[wb.sheetnames[0]]

# Check context around Row 23, Col 15
print("--- Context around Row 23, Col 15 ---")
for r in range(20, 26):
    row_vals = []
    for c in range(13, 17): # roughly around col 15
        row_vals.append(str(ws.cell(row=r, column=c).value))
    print(f"Row {r}: {row_vals}")

# Check context around Row 71, Col 24
print("\n--- Context around Row 71, Col 24 ---")
for r in range(68, 74):
    row_vals = []
    for c in range(22, 26):
        row_vals.append(str(ws.cell(row=r, column=c).value))
    print(f"Row {r}: {row_vals}")
